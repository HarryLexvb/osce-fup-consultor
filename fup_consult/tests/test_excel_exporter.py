"""
Unit tests for Excel exporter.
"""

import io

import pytest
from openpyxl import load_workbook

from fup_consult.exporters.excel_exporter import ExcelExporter
from fup_consult.models import (
    ContratoExperiencia,
    GeneralData,
    OrganoAdministracion,
    ProviderData,
    Representante,
    Socio,
)


@pytest.fixture
def sample_provider_data() -> ProviderData:
    """Provide sample provider data for testing."""
    general = GeneralData(
        ruc="20508238143",
        razon_social="EMPRESA TEST SAC",
        estado="ACTIVO",
        condicion="HABIDO",
        tipo_contribuyente="SOCIEDAD ANONIMA CERRADA",
        departamento="LIMA",
        provincia="LIMA",
        distrito="MIRAFLORES",
        direccion="AV. TEST 123",
        telefonos=["999888777", "998877666"],
        emails=["test@empresa.com"],
    )

    socios = [
        Socio(
            nombre_completo="JUAN PEREZ GARCIA",
            tipo_documento="DNI",
            numero_documento="12345678",
            porcentaje_participacion="50%",
        ),
        Socio(
            nombre_completo="MARIA LOPEZ TORRES",
            tipo_documento="DNI",
            numero_documento="87654321",
            porcentaje_participacion="50%",
        ),
    ]

    representantes = [
        Representante(
            nombre_completo="CARLOS RODRIGUEZ",
            tipo_documento="DNI",
            numero_documento="11223344",
            cargo="GERENTE GENERAL",
        )
    ]

    organos = [
        OrganoAdministracion(
            nombre_completo="ANA MARTINEZ",
            tipo_documento="DNI",
            numero_documento="55667788",
            cargo="DIRECTORA",
        )
    ]

    experiencia = [
        ContratoExperiencia(
            numero_contrato="CTR-2024-001",
            entidad="MINISTERIO DE ECONOMIA",
            objeto_contractual="SERVICIOS DE CONSULTORIA",
            monto=150000.00,
            estado="VIGENTE",
        )
    ]

    return ProviderData(
        general=general,
        socios=socios,
        representantes=representantes,
        organos_administracion=organos,
        experiencia=experiencia,
    )


@pytest.mark.unit
class TestExcelExporter:
    """Test suite for Excel exporter."""

    def test_generate_excel_creates_file(self, sample_provider_data: ProviderData) -> None:
        """Test that Excel file is generated successfully."""
        exporter = ExcelExporter()
        excel_bytes = exporter.generate_excel(sample_provider_data)

        assert isinstance(excel_bytes, bytes)
        assert len(excel_bytes) > 0

    def test_excel_has_all_sheets(self, sample_provider_data: ProviderData) -> None:
        """Test that Excel file contains all required sheets."""
        exporter = ExcelExporter()
        excel_bytes = exporter.generate_excel(sample_provider_data)

        # Load workbook from bytes
        wb = load_workbook(io.BytesIO(excel_bytes))

        expected_sheets = [
            "DatosGenerales",
            "SociosAccionistas",
            "Representantes",
            "OrganosAdministracion",
            "Experiencia",
        ]

        for sheet_name in expected_sheets:
            assert sheet_name in wb.sheetnames

    def test_general_data_sheet_has_correct_content(
        self, sample_provider_data: ProviderData
    ) -> None:
        """Test that general data sheet has correct headers and content."""
        exporter = ExcelExporter()
        excel_bytes = exporter.generate_excel(sample_provider_data)

        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb["DatosGenerales"]

        # Check headers
        assert ws["A1"].value == "Campo"
        assert ws["B1"].value == "Valor"

        # Check RUC is present
        ruc_found = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == "RUC" and row[1] == "20508238143":
                ruc_found = True
                break

        assert ruc_found, "RUC not found in DatosGenerales sheet"

    def test_socios_sheet_has_correct_content(self, sample_provider_data: ProviderData) -> None:
        """Test that socios sheet has correct headers and content."""
        exporter = ExcelExporter()
        excel_bytes = exporter.generate_excel(sample_provider_data)

        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb["SociosAccionistas"]

        # Check headers
        assert ws["A1"].value == "Nombre Completo"
        assert ws["B1"].value == "Tipo Documento"
        assert ws["C1"].value == "Número Documento"
        assert ws["D1"].value == "Participación"

        # Check data rows
        assert ws["A2"].value == "JUAN PEREZ GARCIA"
        assert ws["C2"].value == "12345678"

    def test_representantes_sheet_has_correct_content(
        self, sample_provider_data: ProviderData
    ) -> None:
        """Test that representantes sheet has correct content."""
        exporter = ExcelExporter()
        excel_bytes = exporter.generate_excel(sample_provider_data)

        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb["Representantes"]

        assert ws["A1"].value == "Nombre Completo"
        assert ws["A2"].value == "CARLOS RODRIGUEZ"

    def test_experiencia_sheet_has_correct_content(
        self, sample_provider_data: ProviderData
    ) -> None:
        """Test that experiencia sheet has correct content."""
        exporter = ExcelExporter()
        excel_bytes = exporter.generate_excel(sample_provider_data)

        wb = load_workbook(io.BytesIO(excel_bytes))
        ws = wb["Experiencia"]

        assert ws["A1"].value == "N° Contrato"
        assert ws["A2"].value == "CTR-2024-001"

    def test_empty_sections_handled_gracefully(self) -> None:
        """Test that empty sections are handled without errors."""
        provider_data = ProviderData(
            general=GeneralData(
                ruc="20508238143",
                razon_social="TEST",
                estado="ACTIVO",
                condicion="HABIDO",
                tipo_contribuyente="SAC",
                departamento="LIMA",
                provincia="LIMA",
                distrito="LIMA",
                direccion="DIR",
                telefonos=[],
                emails=[],
            ),
            socios=[],
            representantes=[],
            organos_administracion=[],
            experiencia=[],
        )

        exporter = ExcelExporter()
        excel_bytes = exporter.generate_excel(provider_data)

        assert isinstance(excel_bytes, bytes)
        assert len(excel_bytes) > 0

        # Verify sheets exist even if empty
        wb = load_workbook(io.BytesIO(excel_bytes))
        assert "SociosAccionistas" in wb.sheetnames
