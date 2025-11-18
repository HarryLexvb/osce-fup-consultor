"""
Excel exporter for batch processing results.

This module provides functionality to export multiple provider data
into a single consolidated Excel file with a summary sheet.
"""

import io
import logging
from typing import List, Dict, Any
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelBatchExporter:
    """Exports batch processing results to Excel format."""
    
    def __init__(self) -> None:
        """Initialize Excel batch exporter with styles."""
        self.header_fill = PatternFill(
            start_color="366092",
            end_color="366092",
            fill_type="solid"
        )
        self.summary_fill = PatternFill(
            start_color="4472C4",
            end_color="4472C4",
            fill_type="solid"
        )
        self.success_fill = PatternFill(
            start_color="C6EFCE",
            end_color="C6EFCE",
            fill_type="solid"
        )
        self.header_font = Font(color="FFFFFF", bold=True, size=11)
        self.title_font = Font(bold=True, size=14, color="FFFFFF")
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate_batch_excel(
        self,
        results: List[Dict[str, Any]],
        original_filename: str
    ) -> bytes:
        """
        Generate consolidated Excel file from batch results.
        
        Args:
            results: List of provider data dictionaries
            original_filename: Original input filename for reference
        
        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        
        # Create summary sheet first
        self._create_summary_sheet(wb, results, original_filename)
        
        # Create consolidated data sheet
        self._create_consolidated_sheet(wb, results)
        
        # Create detailed sheets for socios, representantes, and organos
        self._create_socios_detail_sheet(wb, results)
        self._create_representantes_detail_sheet(wb, results)
        self._create_organos_detail_sheet(wb, results)
        
        # Save to bytes
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file.read()
    
    def _create_summary_sheet(
        self,
        wb: Workbook,
        results: List[Dict[str, Any]],
        original_filename: str
    ) -> None:
        """Create summary sheet with statistics and metadata."""
        ws = wb.create_sheet("Resumen", 0)
        
        # Title
        ws.merge_cells('A1:D1')
        title_cell = ws['A1']
        title_cell.value = "RESUMEN DE PROCESAMIENTO BATCH"
        title_cell.fill = self.summary_fill
        title_cell.font = self.title_font
        title_cell.alignment = self.header_alignment
        ws.row_dimensions[1].height = 30
        
        # Metadata section
        row = 3
        metadata = [
            ("📁 Archivo Original", original_filename),
            ("📅 Fecha de Procesamiento", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
            ("📊 Total de RUCs Procesados", len(results)),
            ("", ""),
        ]
        
        for label, value in metadata:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            if label:  # Style label cells
                ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Statistics by estado
        row += 1
        ws[f'A{row}'] = "ESTADÍSTICAS POR ESTADO"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Count by estado
        estado_counts = {}
        for result in results:
            estado = result.get('estado', 'DESCONOCIDO')
            estado_counts[estado] = estado_counts.get(estado, 0) + 1
        
        ws[f'A{row}'] = "Estado"
        ws[f'B{row}'] = "Cantidad"
        ws[f'C{row}'] = "Porcentaje"
        self._apply_header_style(ws, [f'A{row}', f'B{row}', f'C{row}'])
        row += 1
        
        for estado, count in sorted(estado_counts.items()):
            percentage = (count / len(results)) * 100
            ws[f'A{row}'] = estado
            ws[f'B{row}'] = count
            ws[f'C{row}'] = f"{percentage:.1f}%"
            row += 1
        
        # Statistics by tipo_contribuyente
        row += 2
        ws[f'A{row}'] = "ESTADÍSTICAS POR TIPO DE CONTRIBUYENTE"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        tipo_counts = {}
        for result in results:
            tipo = result.get('tipo_contribuyente', 'DESCONOCIDO')
            tipo_counts[tipo] = tipo_counts.get(tipo, 0) + 1
        
        ws[f'A{row}'] = "Tipo de Contribuyente"
        ws[f'B{row}'] = "Cantidad"
        ws[f'C{row}'] = "Porcentaje"
        self._apply_header_style(ws, [f'A{row}', f'B{row}', f'C{row}'])
        row += 1
        
        for tipo, count in sorted(tipo_counts.items(), key=lambda x: x[1], reverse=True)[:10]:
            percentage = (count / len(results)) * 100
            ws[f'A{row}'] = tipo
            ws[f'B{row}'] = count
            ws[f'C{row}'] = f"{percentage:.1f}%"
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
    
    def _create_consolidated_sheet(
        self,
        wb: Workbook,
        results: List[Dict[str, Any]]
    ) -> None:
        """Create consolidated data sheet with all provider information."""
        ws = wb.create_sheet("Datos Consolidados")
        
        # Define headers
        headers = [
            "RUC",
            "Razón Social",
            "Estado",
            "Condición",
            "Tipo de Contribuyente",
            "Domicilio",
            "Departamento",
            "Provincia",
            "Distrito",
            "Teléfonos",
            "Emails",
            "N° Socios",
            "N° Representantes",
            "N° Órganos Administración"
        ]
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
        
        # Write data rows
        for row_idx, result in enumerate(results, start=2):
            ws.cell(row=row_idx, column=1, value=result.get('ruc', ''))
            ws.cell(row=row_idx, column=2, value=result.get('razon_social', ''))
            ws.cell(row=row_idx, column=3, value=result.get('estado', ''))
            ws.cell(row=row_idx, column=4, value=result.get('condicion', ''))
            ws.cell(row=row_idx, column=5, value=result.get('tipo_contribuyente', ''))
            ws.cell(row=row_idx, column=6, value=result.get('domicilio', ''))
            ws.cell(row=row_idx, column=7, value=result.get('departamento', ''))
            ws.cell(row=row_idx, column=8, value=result.get('provincia', ''))
            ws.cell(row=row_idx, column=9, value=result.get('distrito', ''))
            
            # Join lists for display
            telefonos = result.get('telefonos', [])
            ws.cell(row=row_idx, column=10, value=', '.join(telefonos) if telefonos else '')
            
            emails = result.get('emails', [])
            ws.cell(row=row_idx, column=11, value=', '.join(emails) if emails else '')
            
            ws.cell(row=row_idx, column=12, value=result.get('num_socios', 0))
            ws.cell(row=row_idx, column=13, value=result.get('num_representantes', 0))
            ws.cell(row=row_idx, column=14, value=result.get('num_organos', 0))
            
            # Apply borders and success fill for completed items
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = self.thin_border
                if result.get('estado') == 'ACTIVO':
                    cell.fill = self.success_fill
        
        # Adjust column widths
        column_widths = [15, 40, 15, 15, 30, 40, 15, 15, 15, 25, 30, 12, 18, 25]
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Add autofilter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(results) + 1}"
    
    def _apply_header_style(self, ws, cells: List[str]) -> None:
        """Apply header styling to specified cells."""
        for cell_coord in cells:
            cell = ws[cell_coord]
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
    
    def _create_socios_detail_sheet(
        self,
        wb: Workbook,
        results: List[Dict[str, Any]]
    ) -> None:
        """Create detailed sheet for all socios across all companies."""
        ws = wb.create_sheet("Socios Detallados")
        
        # Define headers
        headers = [
            "RUC Empresa",
            "Razón Social Empresa",
            "Nombre Completo Socio",
            "Tipo Doc",
            "Descripción Documento",
            "Número Documento",
            "Participación %",
            "Número de Acciones",
            "Fecha Ingreso"
        ]
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
        
        # Write data rows
        row_idx = 2
        for result in results:
            ruc = result.get('ruc', '')
            razon_social = result.get('razon_social', '')
            socios = result.get('socios', [])
            
            if not socios:
                # Add a row indicating no socios
                ws.cell(row=row_idx, column=1, value=ruc)
                ws.cell(row=row_idx, column=2, value=razon_social)
                ws.cell(row=row_idx, column=3, value='Sin socios registrados')
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col_idx).border = self.thin_border
                row_idx += 1
            else:
                for socio in socios:
                    ws.cell(row=row_idx, column=1, value=ruc)
                    ws.cell(row=row_idx, column=2, value=razon_social)
                    ws.cell(row=row_idx, column=3, value=socio.get('nombre_completo', ''))
                    ws.cell(row=row_idx, column=4, value=socio.get('tipo_documento', ''))
                    ws.cell(row=row_idx, column=5, value=socio.get('desc_tipo_documento', ''))
                    ws.cell(row=row_idx, column=6, value=socio.get('numero_documento', ''))
                    ws.cell(row=row_idx, column=7, value=socio.get('porcentaje_participacion', ''))
                    ws.cell(row=row_idx, column=8, value=socio.get('numero_acciones', ''))
                    ws.cell(row=row_idx, column=9, value=socio.get('fecha_ingreso', ''))
                    
                    # Apply borders
                    for col_idx in range(1, len(headers) + 1):
                        ws.cell(row=row_idx, column=col_idx).border = self.thin_border
                    
                    row_idx += 1
        
        # Adjust column widths
        column_widths = [15, 35, 40, 12, 25, 18, 15, 18, 15]
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Add autofilter
        if row_idx > 2:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_idx - 1}"
    
    def _create_representantes_detail_sheet(
        self,
        wb: Workbook,
        results: List[Dict[str, Any]]
    ) -> None:
        """Create detailed sheet for all representantes across all companies."""
        ws = wb.create_sheet("Representantes Detallados")
        
        # Define headers
        headers = [
            "RUC Empresa",
            "Razón Social Empresa",
            "Nombre Completo",
            "Tipo Doc",
            "Descripción Documento",
            "Número Documento",
            "Cargo",
            "Fecha Desde"
        ]
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
        
        # Write data rows
        row_idx = 2
        for result in results:
            ruc = result.get('ruc', '')
            razon_social = result.get('razon_social', '')
            representantes = result.get('representantes', [])
            
            if not representantes:
                # Add a row indicating no representantes
                ws.cell(row=row_idx, column=1, value=ruc)
                ws.cell(row=row_idx, column=2, value=razon_social)
                ws.cell(row=row_idx, column=3, value='Sin representantes registrados')
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col_idx).border = self.thin_border
                row_idx += 1
            else:
                for rep in representantes:
                    ws.cell(row=row_idx, column=1, value=ruc)
                    ws.cell(row=row_idx, column=2, value=razon_social)
                    ws.cell(row=row_idx, column=3, value=rep.get('nombre_completo', ''))
                    ws.cell(row=row_idx, column=4, value=rep.get('tipo_documento', ''))
                    ws.cell(row=row_idx, column=5, value=rep.get('desc_tipo_documento', ''))
                    ws.cell(row=row_idx, column=6, value=rep.get('numero_documento', ''))
                    ws.cell(row=row_idx, column=7, value=rep.get('cargo', ''))
                    ws.cell(row=row_idx, column=8, value=rep.get('fecha_desde', ''))
                    
                    # Apply borders
                    for col_idx in range(1, len(headers) + 1):
                        ws.cell(row=row_idx, column=col_idx).border = self.thin_border
                    
                    row_idx += 1
        
        # Adjust column widths
        column_widths = [15, 35, 40, 12, 25, 18, 30, 15]
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Add autofilter
        if row_idx > 2:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_idx - 1}"
    
    def _create_organos_detail_sheet(
        self,
        wb: Workbook,
        results: List[Dict[str, Any]]
    ) -> None:
        """Create detailed sheet for all organos de administracion across all companies."""
        ws = wb.create_sheet("Organos Administracion")
        
        # Define headers
        headers = [
            "RUC Empresa",
            "Razón Social Empresa",
            "Nombre Completo",
            "Tipo Doc",
            "Descripción Documento",
            "Número Documento",
            "Tipo de Órgano",
            "Cargo",
            "Fecha Desde"
        ]
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.header_alignment
            cell.border = self.thin_border
        
        # Write data rows
        row_idx = 2
        for result in results:
            ruc = result.get('ruc', '')
            razon_social = result.get('razon_social', '')
            organos = result.get('organos_administracion', [])
            
            if not organos:
                # Add a row indicating no organos
                ws.cell(row=row_idx, column=1, value=ruc)
                ws.cell(row=row_idx, column=2, value=razon_social)
                ws.cell(row=row_idx, column=3, value='Sin órganos de administración registrados')
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col_idx).border = self.thin_border
                row_idx += 1
            else:
                for org in organos:
                    ws.cell(row=row_idx, column=1, value=ruc)
                    ws.cell(row=row_idx, column=2, value=razon_social)
                    ws.cell(row=row_idx, column=3, value=org.get('nombre_completo', ''))
                    ws.cell(row=row_idx, column=4, value=org.get('tipo_documento', ''))
                    ws.cell(row=row_idx, column=5, value=org.get('desc_tipo_documento', ''))
                    ws.cell(row=row_idx, column=6, value=org.get('numero_documento', ''))
                    ws.cell(row=row_idx, column=7, value=org.get('tipo_organo', ''))
                    ws.cell(row=row_idx, column=8, value=org.get('cargo', ''))
                    ws.cell(row=row_idx, column=9, value=org.get('fecha_desde', ''))
                    
                    # Apply borders
                    for col_idx in range(1, len(headers) + 1):
                        ws.cell(row=row_idx, column=col_idx).border = self.thin_border
                    
                    row_idx += 1
        
        # Adjust column widths
        column_widths = [15, 35, 40, 12, 25, 18, 20, 30, 15]
        for col_idx, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Add autofilter
        if row_idx > 2:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_idx - 1}"
