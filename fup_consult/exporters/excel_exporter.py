"""
Excel exporter for provider data.

This module provides functionality to export provider data to Excel format
with multiple sheets for different data categories.
"""

import io
import logging
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from fup_consult.models import ProviderData

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Exports provider data to Excel format."""

    def __init__(self) -> None:
        """Initialize Excel exporter."""
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=11)
        self.header_alignment = Alignment(horizontal="center", vertical="center")

    def generate_excel(self, provider_data: ProviderData) -> bytes:
        """
        Generate Excel file from provider data.

        Args:
            provider_data: Complete provider data

        Returns:
            Excel file as bytes
        """
        wb = Workbook()

        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # Create sheets
        self._create_general_data_sheet(wb, provider_data)
        self._create_socios_sheet(wb, provider_data)
        self._create_representantes_sheet(wb, provider_data)
        self._create_organos_sheet(wb, provider_data)
        self._create_experiencia_sheet(wb, provider_data)

        # Save to bytes
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return excel_file.read()

    def _create_general_data_sheet(self, wb: Workbook, provider_data: ProviderData) -> None:
        """Create general data sheet."""
        ws = wb.create_sheet("DatosGenerales")

        # Headers
        ws["A1"] = "Campo"
        ws["B1"] = "Valor"
        self._apply_header_style(ws, ["A1", "B1"])

        # Data rows
        general = provider_data.general
        rows = [
            ("RUC", general.ruc),
            ("Razón Social", general.razon_social),
            ("Estado", general.estado),
            ("Condición", general.condicion),
            ("Tipo de Contribuyente", general.tipo_contribuyente),
            ("Domicilio Completo", general.domicilio or ""),
            ("Departamento", general.departamento or ""),
            ("Provincia", general.provincia or ""),
            ("Distrito", general.distrito or ""),
            ("Personería", general.personeria or ""),
            ("Teléfonos", ", ".join(general.telefonos) if general.telefonos else ""),
            ("Emails", ", ".join(general.emails) if general.emails else ""),
        ]

        if general.fecha_inscripcion:
            rows.append(("Fecha de Inscripción", general.fecha_inscripcion))
        if general.sistema_emision:
            rows.append(("Sistema de Emisión", general.sistema_emision))
        if general.actividad_economica:
            rows.append(("Actividad Económica", general.actividad_economica))

        for idx, (campo, valor) in enumerate(rows, start=2):
            ws[f"A{idx}"] = campo
            ws[f"B{idx}"] = valor

        self._adjust_column_widths(ws)

    def _create_socios_sheet(self, wb: Workbook, provider_data: ProviderData) -> None:
        """Create shareholders sheet."""
        ws = wb.create_sheet("SociosAccionistas")

        # Headers
        headers = [
            "Nombre Completo",
            "Tipo Documento",
            "Descripción Documento",
            "Número Documento",
            "Participación %",
            "Número de Acciones",
            "Fecha Ingreso"
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            self._apply_header_style(ws, [cell.coordinate])

        # Data rows
        for row_idx, socio in enumerate(provider_data.socios, start=2):
            ws.cell(row=row_idx, column=1, value=socio.nombre_completo)
            ws.cell(row=row_idx, column=2, value=socio.tipo_documento)
            ws.cell(row=row_idx, column=3, value=socio.desc_tipo_documento or "")
            ws.cell(row=row_idx, column=4, value=socio.numero_documento)
            ws.cell(row=row_idx, column=5, value=socio.porcentaje_participacion or "")
            ws.cell(row=row_idx, column=6, value=socio.numero_acciones or "")
            ws.cell(row=row_idx, column=7, value=socio.fecha_ingreso or "")

        if not provider_data.socios:
            ws.cell(row=2, column=1, value="Sin información disponible")

        self._adjust_column_widths(ws)

    def _create_representantes_sheet(self, wb: Workbook, provider_data: ProviderData) -> None:
        """Create legal representatives sheet."""
        ws = wb.create_sheet("Representantes")

        # Headers
        headers = [
            "Nombre Completo",
            "Tipo Documento",
            "Descripción Documento",
            "Número Documento",
            "Cargo",
            "Desde"
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            self._apply_header_style(ws, [cell.coordinate])

        # Data rows
        for row_idx, rep in enumerate(provider_data.representantes, start=2):
            ws.cell(row=row_idx, column=1, value=rep.nombre_completo)
            ws.cell(row=row_idx, column=2, value=rep.tipo_documento)
            ws.cell(row=row_idx, column=3, value=rep.desc_tipo_documento or "")
            ws.cell(row=row_idx, column=4, value=rep.numero_documento)
            ws.cell(row=row_idx, column=5, value=rep.cargo or "")
            ws.cell(row=row_idx, column=6, value=rep.fecha_desde or "")

        if not provider_data.representantes:
            ws.cell(row=2, column=1, value="Sin información disponible")

        self._adjust_column_widths(ws)

    def _create_organos_sheet(self, wb: Workbook, provider_data: ProviderData) -> None:
        """Create administrative bodies sheet."""
        ws = wb.create_sheet("OrganosAdministracion")

        # Headers
        headers = [
            "Nombre Completo",
            "Tipo Documento",
            "Descripción Documento",
            "Número Documento",
            "Tipo de Órgano",
            "Cargo",
            "Desde"
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            self._apply_header_style(ws, [cell.coordinate])

        # Data rows
        for row_idx, org in enumerate(provider_data.organos_administracion, start=2):
            ws.cell(row=row_idx, column=1, value=org.nombre_completo)
            ws.cell(row=row_idx, column=2, value=org.tipo_documento)
            ws.cell(row=row_idx, column=3, value=org.desc_tipo_documento or "")
            ws.cell(row=row_idx, column=4, value=org.numero_documento)
            ws.cell(row=row_idx, column=5, value=org.tipo_organo or "")
            ws.cell(row=row_idx, column=6, value=org.cargo)
            ws.cell(row=row_idx, column=7, value=org.fecha_desde or "")

        if not provider_data.organos_administracion:
            ws.cell(row=2, column=1, value="Sin información disponible")

        self._adjust_column_widths(ws)

    def _create_experiencia_sheet(self, wb: Workbook, provider_data: ProviderData) -> None:
        """Create experience/contracts sheet."""
        ws = wb.create_sheet("Experiencia")

        # Headers
        headers = ["N° Contrato", "Entidad", "Objeto Contractual", "Monto", "Fecha", "Estado"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            self._apply_header_style(ws, [cell.coordinate])

        # Data rows
        for row_idx, exp in enumerate(provider_data.experiencia, start=2):
            ws.cell(row=row_idx, column=1, value=exp.numero_contrato)
            ws.cell(row=row_idx, column=2, value=exp.entidad)
            ws.cell(row=row_idx, column=3, value=exp.objeto_contractual)
            ws.cell(row=row_idx, column=4, value=exp.monto or "")
            ws.cell(row=row_idx, column=5, value=exp.fecha_suscripcion or "")
            ws.cell(row=row_idx, column=6, value=exp.estado or "")

        if not provider_data.experiencia:
            ws.cell(row=2, column=1, value="Sin información disponible")

        self._adjust_column_widths(ws)

    def _apply_header_style(self, ws, cells: List[str]) -> None:
        """Apply header styling to specified cells."""
        for cell_coord in cells:
            cell = ws[cell_coord]
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.header_alignment

    def _adjust_column_widths(self, ws) -> None:
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
